from PyQt5.QtWidgets import QMainWindow ,QApplication,QTableWidgetItem,QMessageBox
from PyQt5.uic import loadUi
from PyQt5 import QtWidgets,QtGui
import openpyxl
import sys,os,re
from delete import remove

                                            # class Main #
class MainUi(QMainWindow):
    def __init__(self):

        super(MainUi,self).__init__()
        loadUi("ui\main.ui",self)

        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.actionRecherche.triggered.connect(self.rech_etud)
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
        self.actionSupprimer_un_Livre.triggered.connect(self.supp_liv) 
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_Emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supprim= Supprimer_emprunt()
        widget.addWidget(supprim)
        widget.setCurrentIndex(widget.currentIndex()+1)
        

    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def Modifier_liv(self):
            modifier_liv = Modifier_livre()
            widget.addWidget(modifier_liv)
            widget.setCurrentIndex(widget.currentIndex()+1)    
    def Afficher_liv(self):    
        afficher_liv = Afficher_livre()
        widget.addWidget(afficher_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def supp_liv(self):
        Supprimer_liv = Supprimer_livre()
        widget.addWidget(Supprimer_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)       
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    

    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_etud(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)    
       
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)


                                #((())) Section Etudiant  (((()))) #                                 
class Ajouter_etud(QMainWindow):
    def __init__(self):

        super(Ajouter_etud,self).__init__()
        loadUi("ui\Ajouter_etud.ui",self) 
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)  
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud) 
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech_etud)
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
        self.actionSupprimer.triggered.connect(self.supp_liv) 
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
        self.pushButton.clicked.connect(self.send)   
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_Emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1)   
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)  
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def Modifier_liv(self):
            modifier_liv = Modifier_livre()
            widget.addWidget(modifier_liv)
            widget.setCurrentIndex(widget.currentIndex()+1)    
    def Afficher_liv(self):    
        afficher_liv = Afficher_livre()
        widget.addWidget(afficher_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def supp_liv(self):
        Supprimer_liv = Supprimer_livre()
        widget.addWidget(Supprimer_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("suppresion eleve") 
    def rech_etud(self):
        Rech = recherche_etud()
        widget.addWidget(Rech)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def clear(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        for row in sheet:
            remove(sheet)
        workbook.save(path)       
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)     

    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1) 
    def popup_suc(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText("eleve enregistree avec succes !")
        msg.setIcon(QMessageBox.Information)
       

        x=msg.exec_()
    def popup_erre(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText("errur d information !")
        msg.setWindowTitle("ERREUR")
        msg.setDetailedText("-Numero inscription doit contenir 8 chiffres \n -phone doit etre formee par 8 chiffres\n -numero de rue doit etre numerique \n -email doit contenir @ et .  ")
        
        

        x=msg.exec_()    
    def send(self):
        tito=True
        n_inscri=self.lineEdit.text()
        nom=self.lineEdit_2.text()
        prenom=self.lineEdit_3.text()
        mail=self.lineEdit_5.text()
        phone=self.lineEdit_6.text()
        n_rue=self.lineEdit_7.text()
        nom_rue=self.lineEdit_8.text()
        ville=self.lineEdit_9.text()
        date=self.dateEdit.text()
        section=self.comboBox.currentText()
        niveau_etud=self.comboBox_2.currentText()
        
        match = re.match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', mail)
        if(len(n_inscri)!=8 or n_inscri.isnumeric()==False):
            tito=False
           
        if(len(phone)!=8 or phone.isnumeric()==False):
            tito=False  
          
        if(nom.isalpha()==False or prenom.isalpha()==False or nom_rue.isalpha()==False or ville.isalpha()==False):
            tito=False        
        if (match==None):
            tito=False
        if(n_rue.isnumeric()==False):
            tito=False 
        d=date[6:]    
        if(2023<int(d)<1800):
            tito=False        
        print("---------------")
        print("numero d inscrit: ",n_inscri,"|nom : ",nom,"|prenom :",prenom,"Date :",date,"mail :",mail,
              "phone :",phone,"adresse :",n_rue+" "+nom_rue+" "+ville,"|niveau etude  :",niveau_etud+section)
        print("---------------")
        self.lineEdit.setText('')
        self.lineEdit_2.setText('')
        self.lineEdit_3.setText('')
        self.lineEdit_5.setText('')
        self.lineEdit_6.setText('')
        self.lineEdit_7.setText('')
        self.lineEdit_8.setText('')
        self.lineEdit_9.setText('')
        path="data\data.xlsx"
        if not os.path.exists(path):
            workbook=openpyxl.Workbook()
            sheet=workbook.active
            heading=["N° inscri","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d Etude","Section"]
            sheet.append(heading)
            workbook.save(path)
        if(tito==True):    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            sheet.append([n_inscri,nom,prenom,date,mail,phone,n_rue+" "+nom_rue+" "+ville,niveau_etud,section])   
            self.popup_suc() 
            self.clear()
            workbook.save(path)  
            
        else:
            print("erreur") 
            self.popup_erre()  

                                            # SUPPRIMER ETUDIANT #   


                  
class Supprimer_etud(QMainWindow):
    def __init__(self):

        super(Supprimer_etud,self).__init__()
        loadUi("ui\Supprimer_etudiant.ui",self)  
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen) 
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.pushButton.clicked.connect(self.supprimer_insc)
        self.pushButton_2.clicked.connect(self.supprimer_sec)
        self.pushButton_3.clicked.connect(self.supprimer_niv)
        self.pushButton_4.clicked.connect(self.supprimer_niv_sec)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech_etud)
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
        self.actionSupprimer.triggered.connect(self.supp_liv) 
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)  
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def Modifier_liv(self):
            modifier_liv = Modifier_livre()
            widget.addWidget(modifier_liv)
            widget.setCurrentIndex(widget.currentIndex()+1)    
    def Afficher_liv(self):    
        afficher_liv = Afficher_livre()
        widget.addWidget(afficher_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def supp_liv(self):
        Supprimer_liv = Supprimer_livre()
        widget.addWidget(Supprimer_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")    
    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1) 
    def rech_etud(self):
        Rech = recherche_etud()
        widget.addWidget(Rech)
        widget.setCurrentIndex(widget.currentIndex()+1)     
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def popup_ins(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f"eleve {self.lineEdit.text()} a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()
    def popup_sec(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f"Section {self.comboBox.currentText()} a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()
           
    def popup_niv(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Niveau {self.comboBox_2.currentText()} annee a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()
    def popup_niv_sec(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" {self.comboBox_3.currentText()} {self.comboBox_4.currentText()}  a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()  

    def supprimer_insc(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        for row in range(2,sheet.max_row):
            if(sheet["A"+str(row)].value ==self.lineEdit.text()):
                sheet.delete_rows(row)
                workbook.save(path) 
                print("deleted")
        self.lineEdit.setText('')        
        self.popup_ins()
    def supprimer_sec(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        for row in range(2,sheet.max_row):
            while(sheet["I"+str(row)].value==self.comboBox.currentText()):
                sheet.delete_rows(row)
                print("row deleted") 
        self.popup_sec()        
        workbook.save(path)               
    def supprimer_niv(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            while(sheet["H"+str(row)].value ==self.comboBox_2.currentText()):
                print("row deleted") 
                sheet.delete_rows(row)
        self.popup_niv()              
        workbook.save(path)   
    def supprimer_niv_sec(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            while(sheet["H"+str(row)].value ==self.comboBox_3.currentText() and sheet["I"+str(row)].value==self.comboBox_4.currentText()):
                print("row deleted") 
                sheet.delete_rows(row)  
        self.popup_niv_sec()            
        workbook.save(path)               
            

       


                                            # MODIFIER ETUDIANT #
class Modifier_etudiant(QMainWindow):
    def __init__(self):

        super(Modifier_etudiant,self).__init__()
        loadUi("ui\Modifier_etud.ui",self)   
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud) 
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.pushButton.clicked.connect(self.modifier_tel)
        self.pushButton_2.clicked.connect(self.modifier_mail)
        self.pushButton_3.clicked.connect(self.modifier_adress)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech_etud)
        self.actionRecherche.triggered.connect(self.rech_liv)
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
        self.actionSupprimer.triggered.connect(self.supp_liv) 
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_Emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)  
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def Modifier_liv(self):
            modifier_liv = Modifier_livre()
            widget.addWidget(modifier_liv)
            widget.setCurrentIndex(widget.currentIndex()+1)    
    def Afficher_liv(self):    
        afficher_liv = Afficher_livre()
        widget.addWidget(afficher_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def supp_liv(self):
        Supprimer_liv = Supprimer_livre()
        widget.addWidget(Supprimer_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)  
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")   
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("suppresion eleve")
    def rech_etud(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)     
    def popup_md_tel(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Numero Telephone de  {self.lineEdit_11.text()}  a ete modifiee avec succes !")
        msg.setIcon(QMessageBox.Information)  
        x=msg.exec_()   
    def modifier_tel(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            if(sheet["A"+str(row)].value==self.lineEdit_11.text()):
                sheet["F"+str(row)].value=self.lineEdit.text()
                
                print("done")
        self.lineEdit_11.setText('')  
        self.lineEdit.setText('')         
        self.popup_md_tel() 
        workbook.save(path) 
                                            #modifier email #        
    def popup_md_mail(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Mail de  {self.lineEdit_3.text()}  a ete modifiee avec succes !")
        msg.setIcon(QMessageBox.Information)  
        x=msg.exec_()     
    def modifier_mail(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            if(sheet["A"+str(row)].value==self.lineEdit_3.text()):
                sheet["E"+str(row)].value=self.lineEdit_2.text()
                print("done")
        self.popup_md_mail() 
        self.lineEdit_3.setText('') 
        self.lineEdit_2.setText('') 
        workbook.save(path)  
                                        #Adresse modification#
    def popup_adress(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Adresse de  {self.lineEdit_10.text()}  a ete modifiee avec succes !")
        msg.setIcon(QMessageBox.Information)  
        x=msg.exec_()     
    def modifier_adress(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            if(sheet["A"+str(row)].value==self.lineEdit_10.text()):
                sheet["G"+str(row)].value=self.lineEdit_7.text()+self.lineEdit_8.text()+self.lineEdit_9.text()
                
                print("done")
        self.popup_adress() 
        self.lineEdit_10.setText('') 
        self.lineEdit_7.setText('') 
        self.lineEdit_8.setText('') 
        self.lineEdit_9.setText('') 
        workbook.save(path)  
               
class recherche_etud(QMainWindow):
    def __init__(self):
        super(recherche_etud,self).__init__()
        loadUi("ui\Recherche_etud.ui",self)
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud) 
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.pushButton_2.clicked.connect(self.load_inscri)
        self.pushButton_3.clicked.connect(self.load_Nom)
        self.pushButton_4.clicked.connect(self.load_Prenom)
        self.pushButton_5.clicked.connect(self.load_Section)
        self.pushButton_6.clicked.connect(self.load_Niveau)
        self.pushButton.clicked.connect(self.load_Tel)
        self.pushButton_7.clicked.connect(self.load_Mail)
        self.pushButton_8.clicked.connect(self.load_Adress)
        self.actionRecherche.triggered.connect(self.rech_liv)
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
        self.actionSupprimer.triggered.connect(self.supp_liv) 
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_Emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)

    
    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)  
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def Modifier_liv(self):
            modifier_liv = Modifier_livre()
            widget.addWidget(modifier_liv)
            widget.setCurrentIndex(widget.currentIndex()+1)    
    def Afficher_liv(self):    
        afficher_liv = Afficher_livre()
        widget.addWidget(afficher_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def supp_liv(self):
        Supprimer_liv = Supprimer_livre()
        widget.addWidget(Supprimer_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("suppresion eleve") 
    def load_inscri(self):
        n_inscri=self.lineEdit_2.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==n_inscri:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1  
    def load_Nom(self):
        nom=self.lineEdit_3.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1         
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==nom:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1     

                    row_ind+=1 
                  
                                   
    def load_Prenom(self):
        prenom=self.lineEdit_4.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==prenom:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1                               
    def load_Section(self):
        section=self.lineEdit_5.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==section:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1    
    def load_Niveau(self):
        niv=self.lineEdit_6.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==niv:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1                           
    def load_Tel(self):
        tel=self.lineEdit.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==tel:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1    
    def load_Mail(self):
        mail=self.lineEdit_7.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==mail:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1   
    def load_Adress(self):
        adr=self.lineEdit_8.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==adr:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1                                      
    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)   

class Afficher_etudiants(QMainWindow):
    def __init__(self):
        super(Afficher_etudiants,self).__init__()
        loadUi("ui\Afficher_etud.ui",self)
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech_etud)
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
        self.actionSupprimer.triggered.connect(self.supp_liv) 
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.load_data()
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_Emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def Modifier_liv(self):
            Modifier_liv = Modifier_livre()
            widget.addWidget(Modifier_liv)
            widget.setCurrentIndex(widget.currentIndex()+1)    
    def Afficher_liv(self):    
        afficher_liv = Afficher_livre()
        widget.addWidget(afficher_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def supp_liv(self):
        Supprimer_liv = Supprimer_livre()
        widget.addWidget(Supprimer_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
       
    def rech_etud(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)        
    
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
          
    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)   
   
    
    def load_data(self):
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        print(list_value)
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
             

                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                col_ind+=1                         
            row_ind+=1                            
           
                                   

                                     


                    # (((()))) Section Livre (((()))) #                              
class Ajouter_livre(QMainWindow):
    def __init__(self):

        super(Ajouter_livre,self).__init__()
        loadUi("ui\Ajouter_livre.ui",self) 
        self.actionSupprimer.triggered.connect(self.supp_liv)  
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv) 
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.pushButton.clicked.connect(self.send)   
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.actionRecherche.triggered.connect(self.rech_etud)
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_Emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)   
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    

    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_etud(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1) 
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)       
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def supp_liv(self):
        supprimer_livres = Supprimer_livre()
        widget.addWidget(supprimer_livres)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("suppresion livre") 
    def rech(self):
        Rech = recherche_livre()
        widget.addWidget(Rech)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def clear(self):
        path="data\Books.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        for row in sheet:
            remove(sheet)
        workbook.save(path)       
    def Afficher_liv(self):    
        afficher_livres = Afficher_livre()
        widget.addWidget(afficher_livres)
        widget.setCurrentIndex(widget.currentIndex()+1)     

    def Modifier_liv(self):
        Modifier_liv = Modifier_livre()
        widget.addWidget(Modifier_liv)
        widget.setCurrentIndex(widget.currentIndex()+1) 
    def popup_suc(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText("Livre enregistree avec succes !")
        msg.setIcon(QMessageBox.Information)
       

        x=msg.exec_()
    def popup_erre(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText("errur d information !")
        msg.setWindowTitle("ERREUR")
        msg.setDetailedText("-Referance doit contenir 6 chiffres \n -ISBN doit etre formee par 13 chiffres\n -nombre de page doit etre numerique \n   ")
        
        

        x=msg.exec_()    
    def send(self):
        tito=True
        referance=self.lineEdit.text()
        Titre=self.lineEdit_2.text()
        Auteur=self.lineEdit_3.text()
        nb_Examp=self.lineEdit_5.text()
        nb_Pages=self.lineEdit_6.text()
        isbn=self.lineEdit_10.text()
        
    
        if(len(referance)!=6 or referance.isnumeric()==False):
            tito=False
           
        if(len(isbn)!=13 or isbn.isnumeric()==False):
            tito=False  
          
        if(Titre.isalpha()==False or Auteur.isalpha()==False):
            tito=False        
        if(nb_Examp.isnumeric()==False):
            tito=False 
        if(nb_Pages.isnumeric()==False):
            tito=False          
        print("---------------")
        print("referance: ",referance,"|titre : ",Titre,"|Auteur:",Auteur,"ISBN :",isbn,"nb pages :",nb_Pages,
              "nb examplaire :",nb_Examp)
        print("---------------")
        self.lineEdit.setText('')
        self.lineEdit_2.setText('')
        self.lineEdit_3.setText('')
        self.lineEdit_5.setText('')
        self.lineEdit_6.setText('')
        self.lineEdit_10.setText('')
    
        path="data\Books.xlsx"
        if not os.path.exists(path):
            workbook=openpyxl.Workbook()
            sheet=workbook.active
            heading=["Referance","Titre","Auteur","Num Examplaire","nb Pages","ISBN"]
            sheet.append(heading)
            workbook.save(path)
        if(tito==True):    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            sheet.append([referance,Titre,Auteur,nb_Examp,nb_Pages,isbn])   
            self.popup_suc() 
            self.clear()
            workbook.save(path)  
            
        else:
            print("erreur") 
            self.popup_erre()
class Supprimer_livre(QMainWindow):
    def __init__(self):

        super(Supprimer_livre,self).__init__()
        loadUi("ui\Supprimer_livre.ui",self)  
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen) 
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.pushButton.clicked.connect(self.supprimer_ref)
        self.pushButton_2.clicked.connect(self.supprimer_isbn)
        self.pushButton_3.clicked.connect(self.supprimer_titre)
        self.pushButton_4.clicked.connect(self.supprimer_auteur)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech_etud)
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    

    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_etud(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def Modifier_liv(self):
        Modifier_liv = Modifier_livre()
        widget.addWidget(Modifier_liv)
        widget.setCurrentIndex(widget.currentIndex()+1) 
    def rech_liv(self):
        Rech = recherche_livre()
        widget.addWidget(Rech)
        widget.setCurrentIndex(widget.currentIndex()+1)     
    def Afficher_liv(self):    
        afficher_liv = Afficher_livre()
        widget.addWidget(afficher_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def popup_ref(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f"livre n° {self.lineEdit.text()} a ete supprimee avec succes !")
    
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()
    def popup_isbn(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f"ISBN {self.lineEdit_2.text()} a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()
           
    def popup_Titre(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Titre {self.lineEdit_3.text()}  a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()
    def popup_Auteur(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" {self.lineEdit_4.text()}  a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()  

    def supprimer_ref(self):
        path="data\Books.3.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        for row in range(2,sheet.max_row):
            if(sheet["A"+str(row)].value ==self.lineEdit.text()):
                sheet.delete_rows(row)
                workbook.save(path) 
                print("deleted")
        self.lineEdit.setText('')        
        self.popup_ref()
    def supprimer_isbn(self):
        path="data\Books.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        for row in range(2,sheet.max_row):
            while(sheet["F"+str(row)].value==self.lineEdit_2.text()):
                sheet.delete_rows(row)
                print("row deleted") 
        self.popup_isbn()        
        workbook.save(path)               
    def supprimer_titre(self):
        path="data\Books.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            while(sheet["B"+str(row)].value ==self.lineEdit_3.text()):
                print("row deleted") 
                sheet.delete_rows(row)
        self.popup_Titre()              
        workbook.save(path)   
    def supprimer_auteur(self):
        path="data\Books.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            while(sheet["C"+str(row)].value ==self.lineEdit_4.text()):
                print("row deleted") 
                sheet.delete_rows(row)  
        self.popup_Auteur()            
        workbook.save(path)               
class Modifier_livre(QMainWindow):
    def __init__(self):

        super(Modifier_livre,self).__init__()
        loadUi("ui\Modifier_nb_exmp.ui",self)   
        
        
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech_etud)
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.actionSupprimer.triggered.connect(self.supp_liv) 
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.pushButton.clicked.connect(self.modifier_nb_examp)
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_Emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
        
    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1) 
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)  
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    def rech_etud(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
       
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def Afficher_liv(self):    
        afficher_liv = Afficher_livre()
        widget.addWidget(afficher_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def supp_liv(self):
        Supprimer_liv = Supprimer_livre()
        widget.addWidget(Supprimer_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
 
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)     
    def popup_nb_examp(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Numero d examplaire  {self.lineEdit_3.text()}  a ete modifiee avec succes !")
        msg.setIcon(QMessageBox.Information)  
        x=msg.exec_()   
    def modifier_nb_examp(self):
        path="data\Books.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            if(sheet["A"+str(row)].value==self.lineEdit_3.text()):
                sheet["D"+str(row)].value=self.lineEdit_2.text()
                
                print("done")
        self.lineEdit_3.setText('')  
        self.lineEdit_2.setText('')         
        self.popup_md_tel() 
        workbook.save(path) 


class recherche_livre(QMainWindow):                                                     
        def __init__(self):
            super(recherche_livre,self).__init__()
            loadUi("ui\Recherche_livre.ui",self)
            self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
            self.actionSupprimer.triggered.connect(self.supp_liv) 
            self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
            self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
            self.actionRecherche.triggered.connect(self.rech_etud)
            self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
            self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
            self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
            self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
            self.pushButton_2.clicked.connect(self.load_ref)
            self.pushButton_3.clicked.connect(self.load_titre)
            self.pushButton_4.clicked.connect(self.load_auteur)
            self.pushButton_5.clicked.connect(self.load_examp)
            self.pushButton.clicked.connect(self.load_isbn)
            self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
            self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
            self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
            self.actionModifier_emprunt.triggered.connect(self.modif_emprunt)

        def modif_emprunt(self):
            modif= Modifier_emprunt()
            widget.addWidget(modif)
            widget.setCurrentIndex(widget.currentIndex()+1)

        def supp_emprunt(self):
            supp= Supprimer_emprunt()
            widget.addWidget(supp)
            widget.setCurrentIndex(widget.currentIndex()+1)
        def retour_emprunt(self):
            retour= Retour_emp()
            widget.addWidget(retour)
            widget.setCurrentIndex(widget.currentIndex()+1)
        def ajouter_liv_screen(self):
            Ajouter_liv = Ajouter_livre()
            widget.addWidget(Ajouter_liv)
            widget.setCurrentIndex(widget.currentIndex()+1)
        def supp_liv(self):
            supprimer_livres = Supprimer_livre()
            widget.addWidget(supprimer_livres)
            widget.setCurrentIndex(widget.currentIndex()+1)
            self.statusbar.showMessage("suppresion livre") 
        def ajouter_etud_screen(self):
            Ajouter_etudient = Ajouter_etud()
            widget.addWidget(Ajouter_etudient)
            widget.setCurrentIndex(widget.currentIndex()+1)
            self.statusbar.showMessage("ajouter etud clicked")  
        def supp_etud(self):
            Supprimer_etudiant = Supprimer_etud()
            widget.addWidget(Supprimer_etudiant)
            widget.setCurrentIndex(widget.currentIndex()+1)
            
        
        def Modifier_liv(self):
            Modifier_liv = Modifier_livre()
            widget.addWidget(Modifier_liv)
            widget.setCurrentIndex(widget.currentIndex()+1)
        def Afficher_liv(self):    
            afficher_liv = Afficher_livre()
            widget.addWidget(afficher_liv)
            widget.setCurrentIndex(widget.currentIndex()+1) 
        def Modifier_etud(self):
            Modifier_etudiante = Modifier_etudiant()
            widget.addWidget(Modifier_etudiante)
            widget.setCurrentIndex(widget.currentIndex()+1)
        def rech_etud(self):
            rech = recherche_etud()
            widget.addWidget(rech)
            widget.setCurrentIndex(widget.currentIndex()+1)    
        def Afficher_etud(self):    
            afficher_etudiante = Afficher_etudiants()
            widget.addWidget(afficher_etudiante)
            widget.setCurrentIndex(widget.currentIndex()+1)      
        def supp_liv(self):
                Supprimer_liv = Supprimer_livre()
                widget.addWidget(Supprimer_liv)
                widget.setCurrentIndex(widget.currentIndex()+1)
        def load_isbn(self):
            isbn=self.lineEdit.text()
            self.tableWidget.setColumnCount(9)
            self.tableWidget.setRowCount(200)
            self.tableWidget.setHorizontalHeaderLabels(("Referance ","Titre","Auteur","Num_Examp","Num_Page","ISBN"))
            path="data\Books.xlsx"    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            list_value=list(sheet.values)
            row_ind=0
            for vin in list_value[1:]:
                col_ind=0
                for v in vin:
                    self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                    col_ind+=1                         
                row_ind+=1 
            row_ind=0
            for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
                for cell in row:
                    if cell==isbn:
                        col_ind=0
                        for v in row:
                            self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                            col_ind+=1                         
                        row_ind+=1      
         
        def load_ref(self):
            ref=self.lineEdit_2.text()
            self.tableWidget.setColumnCount(9)
            self.tableWidget.setRowCount(200)
            self.tableWidget.setHorizontalHeaderLabels(("Referance ","Titre","Auteur","Num_Examp","Num_Page","ISBN"))
            path="data\Books.xlsx"    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            list_value=list(sheet.values)
            row_ind=0
            for vtt in list_value[1:]:
                col_ind=0
                for v in vtt:
                    self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                    col_ind+=1                         
                row_ind+=1 
            row_ind=0
            for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
                for cell in row:
                    if cell==ref:
                        col_ind=0
                        for v in row:
                            self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                            col_ind+=1                         
                        row_ind+=1  

        def load_titre(self):
            titre=self.lineEdit_3.text()
            self.tableWidget.setColumnCount(9)
            self.tableWidget.setRowCount(200)
            self.tableWidget.setHorizontalHeaderLabels(("Referance ","Titre","Auteur","Num_Examp","Num_Page","ISBN"))
            path="data\Books.xlsx"    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            list_value=list(sheet.values)
            row_ind=0
            for voo in list_value[1:]:
                col_ind=0
                for v in voo:
                    self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                    col_ind+=1                         
                row_ind+=1 
            row_ind=0
            for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
                for cell in row:
                    if cell==titre:
                        col_ind=0
                        for v in row:
                            self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                            col_ind+=1                         
                        row_ind+=1                                   
         
        def load_auteur(self):
            auteur=self.lineEdit_4.text()
            self.tableWidget.setColumnCount(9)
            self.tableWidget.setRowCount(200)
            self.tableWidget.setHorizontalHeaderLabels(("Referance ","Titre","Auteur","Num_Examp","Num_Page","ISBN"))
            path="data\Books.xlsx"    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            list_value=list(sheet.values)
            row_ind=0
            for ut in list_value[1:]:
                col_ind=0
                for v in ut:
                    self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                    col_ind+=1                         
                row_ind+=1 
            row_ind=0
            for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
                for cell in row:
                    if cell==auteur:
                        col_ind=0
                        for v in row:
                            self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                            col_ind+=1                         
                        row_ind+=1     
        def load_examp(self):
            examp=self.lineEdit_5.text()
            self.tableWidget.setColumnCount(9)
            self.tableWidget.setRowCount(200)
            self.tableWidget.setHorizontalHeaderLabels(("Referance ","Titre","Auteur","Num_Examp","Num_Page","ISBN"))
            path="data\Books.xlsx"    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            list_value=list(sheet.values)
            row_ind=0
            for oo in list_value[1:]:
                col_ind=0
                for v in oo:
                    self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                    col_ind+=1                         
                row_ind+=1 
            row_ind=0
            for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
                for cell in row:
                    if cell==examp:
                        col_ind=0
                        for v in row:
                            self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                            col_ind+=1                         
                        row_ind+=1   
        def ajouter_emprunt_screen(self):
            Ajouter_emp = Ajouter_emprunt()
            widget.addWidget(Ajouter_emp)
            widget.setCurrentIndex(widget.currentIndex()+1)


class Afficher_livre(QMainWindow):
    def __init__(self):
        super(Afficher_livre,self).__init__()
        loadUi("ui\Afficher_livre.ui",self)
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud) 
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech_etud)
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
        self.actionSupprimer.triggered.connect(self.supp_liv) 
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.load_data()
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)        
     
    def supp_liv(self):
        Supprimer_liv = Supprimer_livre()
        widget.addWidget(Supprimer_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    def Modifier_liv(self):
        Modifier_liv = Modifier_livre()
        widget.addWidget(Modifier_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)   
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    def Afficher_etud(self):    
            afficher_etudiante = Afficher_etudiants()
            widget.addWidget(afficher_etudiante)
            widget.setCurrentIndex(widget.currentIndex()+1) 

    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_etud(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)
    
    def load_data(self):
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("Referance ","Titre","Auteur","Num_Examp","Num_Page","ISBN"))
        path="data\Books.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for t in list_value[:]:
            col_ind=0
            for v in t:
                print(v)
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                col_ind+=1                         
            row_ind+=1 

                        # (((()))) Section Emprunt (((()))) #   


class Ajouter_emprunt(QMainWindow):

    def __init__(self):

        super(Ajouter_emprunt,self).__init__()
        loadUi("ui\Ajouter_emprunt.ui",self) 
        self.actionSupprimer.triggered.connect(self.supp_liv)  
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv) 
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.pushButton.clicked.connect(self.send)   
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.actionRecherche.triggered.connect(self.rech_etud)
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_Emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
       
    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    

    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_etud(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1) 
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)       
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def supp_liv(self):
        supprimer_livres = Supprimer_livre()
        widget.addWidget(supprimer_livres)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("suppresion livre") 
    def rech(self):
        Rech = recherche_livre()
        widget.addWidget(Rech)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def clear(self):
        path="data\Books.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        for row in sheet:
            remove(sheet)
        workbook.save(path)       
    def Afficher_liv(self):    
        afficher_livres = Afficher_livre()
        widget.addWidget(afficher_livres)
        widget.setCurrentIndex(widget.currentIndex()+1)     

    def Modifier_liv(self):
        Modifier_liv = Modifier_livre()
        widget.addWidget(Modifier_liv)
        widget.setCurrentIndex(widget.currentIndex()+1) 
    def popup_suc(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText("Livre enregistree avec succes !")
        msg.setIcon(QMessageBox.Information)
       

        x=msg.exec_()
    def popup_erre(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText("errur d information !")
        msg.setWindowTitle("ERREUR")
        msg.setDetailedText("-Referance doit contenir 6 chiffres \n -ISBN doit etre formee par 13 chiffres\n -nombre de page doit etre numerique \n   ")
        
        

        x=msg.exec_()    
    def send(self):
        tito=True
        referance=self.lineEdit.text()
        Titre=self.lineEdit_2.text()
        Auteur=self.lineEdit_3.text()
        nb_Examp=self.lineEdit_5.text()
        nb_Pages=self.lineEdit_6.text()
        isbn=self.lineEdit_10.text()
        
    
        if(len(referance)!=6 or referance.isnumeric()==False):
            tito=False
           
        if(len(isbn)!=13 or isbn.isnumeric()==False):
            tito=False  
          
        if(Titre.isalpha()==False or Auteur.isalpha()==False):
            tito=False        
        if(nb_Examp.isnumeric()==False):
            tito=False 
        if(nb_Pages.isnumeric()==False):
            tito=False          
        print("---------------")
        print("referance: ",referance,"|titre : ",Titre,"|Auteur:",Auteur,"ISBN :",isbn,"nb pages :",nb_Pages,
              "nb examplaire :",nb_Examp)
        print("---------------")
        self.lineEdit.setText('')
        self.lineEdit_2.setText('')
        self.lineEdit_3.setText('')
        self.lineEdit_5.setText('')
        self.lineEdit_6.setText('')
        self.lineEdit_10.setText('')
    
        path="data\Books.xlsx"
        if not os.path.exists(path):
            workbook=openpyxl.Workbook()
            sheet=workbook.active
            heading=["Referance","Titre","Auteur","Num Examplaire","nb Pages","ISBN"]
            sheet.append(heading)
            workbook.save(path)
        if(tito==True):    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            sheet.append([referance,Titre,Auteur,nb_Examp,nb_Pages,isbn])   
            self.popup_suc() 
            self.clear()
            workbook.save(path)  
            
        else:
            print("erreur") 
            self.popup_erre() 

class Retour_emp(QMainWindow):

    def __init__(self):

        super(Retour_emp,self).__init__()
        loadUi("ui\Retour_emprunt.ui",self)   
        
        
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech_etud)
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.actionSupprimer.triggered.connect(self.supp_liv) 
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.pushButton.clicked.connect(self.modifier_nb_examp)
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_Emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
        
      
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)  
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    def rech_etud(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
       
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def Afficher_liv(self):    
        afficher_liv = Afficher_livre()
        widget.addWidget(afficher_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def supp_liv(self):
        Supprimer_liv = Supprimer_livre()
        widget.addWidget(Supprimer_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
 
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)     
    def popup_nb_examp(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Numero d examplaire  {self.lineEdit_3.text()}  a ete modifiee avec succes !")
        msg.setIcon(QMessageBox.Information)  
        x=msg.exec_()   
    def modifier_nb_examp(self):
        path="data\Books.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            if(sheet["A"+str(row)].value==self.lineEdit_3.text()):
                sheet["D"+str(row)].value=self.lineEdit_2.text()
                
                print("done")
        self.lineEdit_3.setText('')  
        self.lineEdit_2.setText('')         
        self.popup_md_tel() 
        workbook.save(path) 

class Supprimer_emprunt(QMainWindow):
    def __init__(self):

        super(Supprimer_emprunt,self).__init__()
        loadUi("ui\Supprimer_emprunt.ui",self)  
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen) 
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.pushButton.clicked.connect(self.supprimer_ref)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech_etud)
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionSupprimer.triggered.connect(self.supp_liv) 
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.actionModifier_emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)
       
    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def supp_liv(self):
        supprimer_livres = Supprimer_livre()
        widget.addWidget(supprimer_livres)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("suppresion livre") 
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    

    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_etud(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def Modifier_liv(self):
        Modifier_liv = Modifier_livre()
        widget.addWidget(Modifier_liv)
        widget.setCurrentIndex(widget.currentIndex()+1) 
    def rech_liv(self):
        Rech = recherche_livre()
        widget.addWidget(Rech)
        widget.setCurrentIndex(widget.currentIndex()+1)     
    def Afficher_liv(self):    
        afficher_liv = Afficher_livre()
        widget.addWidget(afficher_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def popup_ref(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f"livre n° {self.lineEdit.text()} a ete supprimee avec succes !")
    
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()
    def popup_isbn(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f"ISBN {self.lineEdit_2.text()} a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()
           
    def popup_Titre(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Titre {self.lineEdit_3.text()}  a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()
    def popup_Auteur(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" {self.lineEdit_4.text()}  a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()  

    def supprimer_ref(self):
        path="data\Books.3.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        for row in range(2,sheet.max_row):
            if(sheet["A"+str(row)].value ==self.lineEdit.text()):
                sheet.delete_rows(row)
                workbook.save(path) 
                print("deleted")
        self.lineEdit.setText('')        
        self.popup_ref()

class Modifier_emprunt(QMainWindow):
    def __init__(self):

        super(Modifier_emprunt,self).__init__()
        loadUi("ui\Modifier_emprunt.ui",self)   
        
        
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech_etud)
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.actionSupprimer.triggered.connect(self.supp_liv) 
        self.actionRecherche.triggered.connect(self.rech_liv)
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
       

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def Modifier_liv(self):
        Modifier_liv = Modifier_livre()
        widget.addWidget(Modifier_liv)
        widget.setCurrentIndex(widget.currentIndex()+1) 
   
        
    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1) 
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)  
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    def rech_etud(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
       
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def Afficher_liv(self):    
        afficher_liv = Afficher_livre()
        widget.addWidget(afficher_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def supp_liv(self):
        Supprimer_liv = Supprimer_livre()
        widget.addWidget(Supprimer_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
 
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)     
    def popup_nb_examp(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Numero d examplaire  {self.lineEdit_3.text()}  a ete modifiee avec succes !")
        msg.setIcon(QMessageBox.Information)  
        x=msg.exec_()   

class recherche_emprunt(QMainWindow):                                                     
        def __init__(self):
            super(recherche_emprunt,self).__init__()
            loadUi("ui\Recherche_emprunt.ui",self)
            self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
            self.actionSupprimer.triggered.connect(self.supp_liv) 
            self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
            self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
            self.actionRecherche.triggered.connect(self.rech_etud)
            self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
            self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
            self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
            self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
            self.pushButton_2.clicked.connect(self.load_ref)
            self.pushButton_3.clicked.connect(self.load_titre)
            self.pushButton_4.clicked.connect(self.load_auteur)
            self.pushButton_5.clicked.connect(self.load_examp)
            self.pushButton.clicked.connect(self.load_isbn)
            self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
            self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
            self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
            self.actionModifier_emprunt.triggered.connect(self.modif_emprunt)

        def modif_emprunt(self):
            modif= Modifier_emprunt()
            widget.addWidget(modif)
            widget.setCurrentIndex(widget.currentIndex()+1)

        def supp_emprunt(self):
            supp= Supprimer_emprunt()
            widget.addWidget(supp)
            widget.setCurrentIndex(widget.currentIndex()+1)
        def retour_emprunt(self):
            retour= Retour_emp()
            widget.addWidget(retour)
            widget.setCurrentIndex(widget.currentIndex()+1)
        def ajouter_liv_screen(self):
            Ajouter_liv = Ajouter_livre()
            widget.addWidget(Ajouter_liv)
            widget.setCurrentIndex(widget.currentIndex()+1)
        def supp_liv(self):
            supprimer_livres = Supprimer_livre()
            widget.addWidget(supprimer_livres)
            widget.setCurrentIndex(widget.currentIndex()+1)
            self.statusbar.showMessage("suppresion livre") 
        def ajouter_etud_screen(self):
            Ajouter_etudient = Ajouter_etud()
            widget.addWidget(Ajouter_etudient)
            widget.setCurrentIndex(widget.currentIndex()+1)
            self.statusbar.showMessage("ajouter etud clicked")  
        def supp_etud(self):
            Supprimer_etudiant = Supprimer_etud()
            widget.addWidget(Supprimer_etudiant)
            widget.setCurrentIndex(widget.currentIndex()+1)
            
        
        def Modifier_liv(self):
            Modifier_liv = Modifier_livre()
            widget.addWidget(Modifier_liv)
            widget.setCurrentIndex(widget.currentIndex()+1)
        def Afficher_liv(self):    
            afficher_liv = Afficher_livre()
            widget.addWidget(afficher_liv)
            widget.setCurrentIndex(widget.currentIndex()+1) 
        def Modifier_etud(self):
            Modifier_etudiante = Modifier_etudiant()
            widget.addWidget(Modifier_etudiante)
            widget.setCurrentIndex(widget.currentIndex()+1)
        def rech_etud(self):
            rech = recherche_etud()
            widget.addWidget(rech)
            widget.setCurrentIndex(widget.currentIndex()+1)   
        def ajouter_emprunt_screen(self):
            Ajouter_emp = Ajouter_emprunt()
            widget.addWidget(Ajouter_emp)
            widget.setCurrentIndex(widget.currentIndex()+1)       
        def Afficher_etud(self):    
            afficher_etudiante = Afficher_etudiants()
            widget.addWidget(afficher_etudiante)
            widget.setCurrentIndex(widget.currentIndex()+1)      
        def supp_liv(self):
                Supprimer_liv = Supprimer_livre()
                widget.addWidget(Supprimer_liv)
                widget.setCurrentIndex(widget.currentIndex()+1)
        def load_isbn(self):
            isbn=self.lineEdit.text()
            self.tableWidget.setColumnCount(9)
            self.tableWidget.setRowCount(200)
            self.tableWidget.setHorizontalHeaderLabels(("Referance ","Titre","Auteur","Num_Examp","Num_Page","ISBN"))
            path="data\Books.xlsx"    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            list_value=list(sheet.values)
            row_ind=0
            for vin in list_value[1:]:
                col_ind=0
                for v in vin:
                    self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                    col_ind+=1                         
                row_ind+=1 
            row_ind=0
            for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
                for cell in row:
                    if cell==isbn:
                        col_ind=0
                        for v in row:
                            self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                            col_ind+=1                         
                        row_ind+=1      
         
        def load_ref(self):
            ref=self.lineEdit_2.text()
            self.tableWidget.setColumnCount(9)
            self.tableWidget.setRowCount(200)
            self.tableWidget.setHorizontalHeaderLabels(("Referance ","Titre","Auteur","Num_Examp","Num_Page","ISBN"))
            path="data\Books.xlsx"    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            list_value=list(sheet.values)
            row_ind=0
            for vtt in list_value[1:]:
                col_ind=0
                for v in vtt:
                    self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                    col_ind+=1                         
                row_ind+=1 
            row_ind=0
            for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
                for cell in row:
                    if cell==ref:
                        col_ind=0
                        for v in row:
                            self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                            col_ind+=1                         
                        row_ind+=1  

        def load_titre(self):
            titre=self.lineEdit_3.text()
            self.tableWidget.setColumnCount(9)
            self.tableWidget.setRowCount(200)
            self.tableWidget.setHorizontalHeaderLabels(("Referance ","Titre","Auteur","Num_Examp","Num_Page","ISBN"))
            path="data\Books.xlsx"    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            list_value=list(sheet.values)
            row_ind=0
            for voo in list_value[1:]:
                col_ind=0
                for v in voo:
                    self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                    col_ind+=1                         
                row_ind+=1 
            row_ind=0
            for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
                for cell in row:
                    if cell==titre:
                        col_ind=0
                        for v in row:
                            self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                            col_ind+=1                         
                        row_ind+=1                                   
         
        def load_auteur(self):
            auteur=self.lineEdit_4.text()
            self.tableWidget.setColumnCount(9)
            self.tableWidget.setRowCount(200)
            self.tableWidget.setHorizontalHeaderLabels(("Referance ","Titre","Auteur","Num_Examp","Num_Page","ISBN"))
            path="data\Books.xlsx"    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            list_value=list(sheet.values)
            row_ind=0
            for ut in list_value[1:]:
                col_ind=0
                for v in ut:
                    self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                    col_ind+=1                         
                row_ind+=1 
            row_ind=0
            for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
                for cell in row:
                    if cell==auteur:
                        col_ind=0
                        for v in row:
                            self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                            col_ind+=1                         
                        row_ind+=1     
        def load_examp(self):
            examp=self.lineEdit_5.text()
            self.tableWidget.setColumnCount(9)
            self.tableWidget.setRowCount(200)
            self.tableWidget.setHorizontalHeaderLabels(("Referance ","Titre","Auteur","Num_Examp","Num_Page","ISBN"))
            path="data\Books.xlsx"    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            list_value=list(sheet.values)
            row_ind=0
            for oo in list_value[1:]:
                col_ind=0
                for v in oo:
                    self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                    col_ind+=1                         
                row_ind+=1 
            row_ind=0
            for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
                for cell in row:
                    if cell==examp:
                        col_ind=0
                        for v in row:
                            self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                            col_ind+=1                         
                        row_ind+=1   
class Afficher_emprunt(QMainWindow):
    def __init__(self):
        super(Afficher_emprunt,self).__init__()
        loadUi("ui\Afficher_emprunt.ui",self)
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech_etud)
        self.actionRecherche_par_Referance.triggered.connect(self.rech_liv)
        self.actionAjouter_un_nouvel_Livre.triggered.connect(self.ajouter_liv_screen)
        self.actionSupprimer.triggered.connect(self.supp_liv) 
        self.actionContenue_du_dictionnaire_livre.triggered.connect(self.Afficher_liv)
        self.actionModifier_nombre_d_exemplaires_d_un_Livre.triggered.connect(self.Modifier_liv)
        self.actionAjouter_un_nouvel_emprunt.triggered.connect(self.ajouter_emprunt_screen)
        self.actionRetour_d_un_emprunt.triggered.connect(self.retour_emprunt)
        self.load_data()
        self.actionSupprimer_d_un_emprunt.triggered.connect(self.supp_emprunt)
        self.actionModifier_Emprunt.triggered.connect(self.modif_emprunt)

    def modif_emprunt(self):
        modif= Modifier_emprunt()
        widget.addWidget(modif)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def supp_emprunt(self):
        supp= Supprimer_emprunt()
        widget.addWidget(supp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def retour_emprunt(self):
        retour= Retour_emp()
        widget.addWidget(retour)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_emprunt_screen(self):
        Ajouter_emp = Ajouter_emprunt()
        widget.addWidget(Ajouter_emp)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def ajouter_liv_screen(self):
        Ajouter_liv = Ajouter_livre()
        widget.addWidget(Ajouter_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def Modifier_liv(self):
            Modifier_liv = Modifier_livre()
            widget.addWidget(Modifier_liv)
            widget.setCurrentIndex(widget.currentIndex()+1)    
    def Afficher_liv(self):    
        afficher_liv = Afficher_livre()
        widget.addWidget(afficher_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def supp_liv(self):
        Supprimer_liv = Supprimer_livre()
        widget.addWidget(Supprimer_liv)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech_liv(self):
        rech = recherche_livre()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
       
    def rech_etud(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)        
    
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
          
    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)   
   
    
    def load_data(self):  
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        print(list_value)
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
             

                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                col_ind+=1                         
            row_ind+=1                                                                
if __name__=='__main__':
     app=QApplication(sys.argv)
     ui=MainUi()
     MainWindow=MainUi()
     widget=QtWidgets.QStackedWidget()
     widget.addWidget(MainWindow)
     widget.setWindowTitle('ISIMM Library')
     widget.setWindowIcon(QtGui.QIcon('ui\images\lg2.png'))
     widget.setFixedHeight(600)
     widget.setFixedWidth(1250)
     widget.show()
     app.exec_()            